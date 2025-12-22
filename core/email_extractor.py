"""Extractor de adjuntos de correos de Outlook con filtrado dual y auditoría."""

import gc
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import List

import pythoncom
import win32com.client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.backend_base import (
    BackendBase,
    FaseProceso as FaseBase,
    NivelMensaje,
    EstadoProceso,
    EstadisticasBase
)
from utils.date_handler import normalize_to_naive, get_day_start, get_day_end
from utils.audit_mails import EmailAuditor


class FaseProceso(Enum):
    """Fases específicas del proceso de extracción"""
    INICIAL = "inicial"
    FILTRADO = "filtrado"
    DESCARGA = "descarga"
    FINALIZACION = "finalizacion"


@dataclass
class EstadisticasExtraccion(EstadisticasBase):
    """Estadísticas del proceso de extracción"""
    total_correos: int = 0
    correos_procesados: int = 0
    adjuntos_descargados: int = 0
    adjuntos_fallidos: int = 0
    tamaño_total_mb: float = 0.0
    archivos_descargados: List[dict] = field(default_factory=list)
    
    @property
    def tasa_exito(self) -> float:
        """Porcentaje de éxito"""
        total = self.adjuntos_descargados + self.adjuntos_fallidos
        return (self.adjuntos_descargados / total * 100) if total > 0 else 0.0


class ExtractorAdjuntosOutlook(BackendBase):
    """
    Extractor de adjuntos de Outlook con filtrado dual (asunto + nombre de archivo).
    
    Funcionalidades: filtrado por frases, rango de fechas, early exit,
    logging detallado, auditoría completa y generación de reportes Excel.
    """
    
    def __init__(self, 
                 callback_mensaje=None,
                 callback_progreso=None, 
                 callback_estado=None):
        """Inicializa el extractor"""
        super().__init__(callback_mensaje, callback_progreso, callback_estado)
        
        self.estadisticas = EstadisticasExtraccion()
        self.outlook = None
        self.namespace = None
        self.auditor = None  # Se inicializa en extraer_adjuntos
        self._cancelado = False  # Flag para cancelación
        
        self.config = {
            "max_reintentos": 3,
            "liberar_memoria_cada": 1000,
            "solo_cache_local": True
        }
    
    def validar_parametros(self, frases: List[str], destino: str, 
                          outlook_folder: str, fecha_inicio: datetime, 
                          fecha_fin: datetime) -> tuple[bool, str]:
        """Valida los parámetros de extracción"""
        
        # Nota: frases puede estar vacío (procesar todos los correos)
        
        if not destino or not destino.strip():
            return False, "Debe seleccionar una carpeta de destino"
        
        if not outlook_folder or not outlook_folder.strip():
            return False, "Debe seleccionar una bandeja de correo"
        
        es_valido, mensaje = self._validar_rango_fechas(fecha_inicio, fecha_fin)
        if not es_valido:
            return False, mensaje
        
        dias_diferencia = (fecha_fin - fecha_inicio).days
        if dias_diferencia > 365:
            return False, "El rango de fechas no puede superar 1 año"
        
        return True, ""
    
    def _procesar_principal(self, *args, **kwargs) -> dict:
        """Delegado al método extraer_adjuntos"""
        return {}
    
    def _generar_reporte(self) -> dict:
        """Genera reporte final"""
        return self._generar_resultado()
    
    def extraer_adjuntos(self, frases: List[str], destino: str, 
                        outlook_folder: str, fecha_inicio: datetime, 
                        fecha_fin: datetime) -> dict:
        """
        Extrae adjuntos de correos según criterios.
        
        Args:
            frases: Frases para filtrar (si está vacío, procesa todos)
            destino: Carpeta destino
            outlook_folder: Ruta carpeta Outlook
            fecha_inicio: Fecha inicial
            fecha_fin: Fecha final
            
        Returns:
            dict: Estadísticas del proceso
        """
        
        self.fase_actual = FaseProceso.INICIAL
        self.estadisticas.tiempo_inicio = datetime.now()
        
        # Inicializar auditor
        self.auditor = EmailAuditor(ruta_salida=destino)
        
        try:
            es_valido, mensaje = self.validar_parametros(
                frases, destino, outlook_folder, fecha_inicio, fecha_fin
            )
            if not es_valido:
                raise ValueError(mensaje)
            
            self._conectar_outlook()
            carpeta = self._obtener_carpeta(outlook_folder)
            
            correos_data = self._filtrar_correos_optimizado(
                carpeta, frases, fecha_inicio, fecha_fin
            )
            
            if not correos_data:
                self._enviar_mensaje(
                    FaseProceso.FILTRADO,
                    NivelMensaje.WARNING,
                    "No se encontraron correos"
                )
                # Exportar auditoría aunque esté vacía
                self._exportar_auditoria()
                return self._generar_resultado_vacio()
            
            self._descargar_adjuntos(correos_data, carpeta, destino, frases)
            self._generar_excel_listado(destino)
            
            # Exportar auditoría
            self._exportar_auditoria()
            
            self.estadisticas.tiempo_fin = datetime.now()
            resultado = self._generar_resultado()
            
            self.logger.escribir_estadisticas(resultado)
            return resultado
            
        except Exception as e:
            self._cambiar_estado(EstadoProceso.ERROR)
            self._enviar_mensaje(
                self.fase_actual,
                NivelMensaje.ERROR,
                f"Error: {str(e)}"
            )
            self.logger.error(f"Error en extracción: {str(e)}", exc_info=True)
            raise
        finally:
            self._desconectar_outlook()
    
    def _conectar_outlook(self):
        """Conecta con Outlook"""
        self._enviar_mensaje(
            FaseProceso.INICIAL,
            NivelMensaje.INFO,
            "Conectando con Outlook..."
        )
        
        try:
            pythoncom.CoInitialize()
            self.outlook = win32com.client.Dispatch("Outlook.Application")
            self.namespace = self.outlook.GetNamespace("MAPI")
            
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.SUCCESS,
                "Conexión exitosa"
            )
        except Exception as e:
            raise RuntimeError(f"Error al conectar: {str(e)}")
    
    def _desconectar_outlook(self):
        """Desconecta de Outlook"""
        try:
            if self.namespace:
                self.namespace = None
            if self.outlook:
                self.outlook = None
            pythoncom.CoUninitialize()
        except:
            pass
    
    def _obtener_carpeta(self, outlook_folder: str):
        """Obtiene carpeta de Outlook"""
        self._enviar_mensaje(
            FaseProceso.INICIAL,
            NivelMensaje.INFO,
            f"Buscando carpeta: {outlook_folder}"
        )
        
        try:
            partes = outlook_folder.split("\\")
            carpeta_actual = None
            
            for folder in self.namespace.Folders:
                if folder.Name == partes[0]:
                    carpeta_actual = folder
                    break
            
            if not carpeta_actual:
                raise ValueError(f"Cuenta no encontrada: {partes[0]}")
            
            for parte in partes[1:]:
                encontrada = False
                for subfolder in carpeta_actual.Folders:
                    if subfolder.Name == parte:
                        carpeta_actual = subfolder
                        encontrada = True
                        break
                
                if not encontrada:
                    raise ValueError(f"Carpeta no encontrada: {parte}")
            
            self._enviar_mensaje(
                FaseProceso.INICIAL,
                NivelMensaje.SUCCESS,
                f"Carpeta encontrada: {carpeta_actual.Name}"
            )
            
            return carpeta_actual
            
        except Exception as e:
            raise RuntimeError(f"Error al obtener carpeta: {str(e)}")
    
    def _filtrar_correos_optimizado(self, carpeta, frases: List[str], 
                                    fecha_inicio: datetime, fecha_fin: datetime) -> List[dict]:
        """
        Filtra correos con early exit, normalización de fechas y auditoría completa.
        
        Args:
            carpeta: Carpeta de Outlook
            frases: Lista de frases para filtrar (vacío = todos)
            fecha_inicio: Fecha inicial
            fecha_fin: Fecha final
            
        Returns:
            List[dict]: Correos filtrados con índice, fecha y asunto
        """
        
        self._cambiar_fase(FaseProceso.FILTRADO)
        self._cambiar_estado(EstadoProceso.FILTRANDO)
        
        # Normalizar fechas a naive y cubrir días completos
        fecha_inicio_norm = get_day_start(fecha_inicio)
        fecha_fin_norm = get_day_end(fecha_fin)
        
        # Normalizar frases (si están vacías, modo sin filtro)
        frases_norm = [f.strip().lower() for f in frases if f.strip()]
        modo_sin_filtro = len(frases_norm) == 0
        
        if modo_sin_filtro:
            self._enviar_mensaje(
                FaseProceso.FILTRADO,
                NivelMensaje.INFO,
                "⚠️ Modo sin filtro de frases: procesando TODOS los correos"
            )
        
        items = carpeta.Items
        items.Sort("[ReceivedTime]", True)
        
        total_items = items.Count
        correos_filtrados = []
        fuera_rango_consecutivos = 0
        
        self._enviar_mensaje(
            FaseProceso.FILTRADO,
            NivelMensaje.INFO,
            f"Escaneando {total_items} correos..."
        )
        
        for idx in range(1, total_items + 1):
            try:
                item = items.Item(idx)
                
                # Extraer metadata
                try:
                    entry_id = item.EntryID
                except:
                    entry_id = f"UNKNOWN_{idx}"
                
                fecha_correo = normalize_to_naive(item.ReceivedTime)
                asunto = str(item.Subject)
                
                try:
                    remitente = str(item.SenderEmailAddress)
                except:
                    remitente = "UNKNOWN"
                
                tiene_adjuntos = item.Attachments.Count > 0
                num_adjuntos = item.Attachments.Count
                
                # Obtener nombres de adjuntos
                adjuntos_nombres = []
                if tiene_adjuntos:
                    try:
                        for adjunto in item.Attachments:
                            adjuntos_nombres.append(adjunto.FileName)
                    except:
                        adjuntos_nombres = ["ERROR_READING_ATTACHMENTS"]
                
                # Validar fecha
                cumple_fecha = fecha_inicio_norm <= fecha_correo <= fecha_fin_norm
                
                # Early exit: si encontramos 100 correos consecutivos fuera del rango, salir
                if fecha_correo < fecha_inicio_norm:
                    fuera_rango_consecutivos += 1
                    
                    # Auditar correo rechazado por fecha
                    self.auditor.registrar_correo(
                        entry_id=entry_id,
                        received_time=fecha_correo,
                        subject=asunto,
                        sender=remitente,
                        cumple_fecha=False,
                        cumple_frases=False,  # No se evalúa
                        tiene_adjuntos=tiene_adjuntos,
                        num_adjuntos=num_adjuntos,
                        adjuntos_nombres=adjuntos_nombres,
                        adjuntos_descargados=0,
                        estado_final="RECHAZADO",
                        motivo_rechazo="Fuera de rango de fechas (anterior al inicio)",
                        fase_proceso="FILTRADO"
                    )
                    
                    if fuera_rango_consecutivos >= 100:
                        self._enviar_mensaje(
                            FaseProceso.FILTRADO,
                            NivelMensaje.INFO,
                            f"Early exit: 100 correos consecutivos fuera de rango"
                        )
                        del item
                        break
                    del item
                    continue
                
                if fecha_correo > fecha_fin_norm:
                    # Auditar correo rechazado por fecha
                    self.auditor.registrar_correo(
                        entry_id=entry_id,
                        received_time=fecha_correo,
                        subject=asunto,
                        sender=remitente,
                        cumple_fecha=False,
                        cumple_frases=False,  # No se evalúa
                        tiene_adjuntos=tiene_adjuntos,
                        num_adjuntos=num_adjuntos,
                        adjuntos_nombres=adjuntos_nombres,
                        adjuntos_descargados=0,
                        estado_final="RECHAZADO",
                        motivo_rechazo="Fuera de rango de fechas (posterior al fin)",
                        fase_proceso="FILTRADO"
                    )
                    
                    del item
                    continue
                
                fuera_rango_consecutivos = 0
                
                # Validar frases
                cumple_frases = False
                if modo_sin_filtro:
                    cumple_frases = True
                else:
                    asunto_lower = asunto.lower()
                    cumple_frases = any(frase in asunto_lower for frase in frases_norm)
                
                # Evaluar si se acepta el correo
                if cumple_frases and tiene_adjuntos:
                    # Correo aceptado
                    correos_filtrados.append({
                        'indice': idx,
                        'entry_id': entry_id,
                        'fecha': fecha_correo,
                        'asunto': asunto,
                        'remitente': remitente,
                        'num_adjuntos': num_adjuntos,
                        'adjuntos_nombres': adjuntos_nombres
                    })
                    
                    # Auditar correo aceptado
                    self.auditor.registrar_correo(
                        entry_id=entry_id,
                        received_time=fecha_correo,
                        subject=asunto,
                        sender=remitente,
                        cumple_fecha=True,
                        cumple_frases=True,
                        tiene_adjuntos=True,
                        num_adjuntos=num_adjuntos,
                        adjuntos_nombres=adjuntos_nombres,
                        adjuntos_descargados=0,  # Se actualizará en descarga
                        estado_final="PROCESADO",
                        motivo_rechazo="",
                        fase_proceso="FILTRADO"
                    )
                else:
                    # Correo rechazado
                    motivo = []
                    if not cumple_frases and not modo_sin_filtro:
                        motivo.append("No coincide con frases de búsqueda")
                    if not tiene_adjuntos:
                        motivo.append("No tiene adjuntos")
                    
                    motivo_rechazo = " | ".join(motivo) if motivo else "Filtrado general"
                    
                    self.auditor.registrar_correo(
                        entry_id=entry_id,
                        received_time=fecha_correo,
                        subject=asunto,
                        sender=remitente,
                        cumple_fecha=True,
                        cumple_frases=cumple_frases,
                        tiene_adjuntos=tiene_adjuntos,
                        num_adjuntos=num_adjuntos,
                        adjuntos_nombres=adjuntos_nombres,
                        adjuntos_descargados=0,
                        estado_final="RECHAZADO",
                        motivo_rechazo=motivo_rechazo,
                        fase_proceso="FILTRADO"
                    )
                
                del item
                
                # Verificar cancelación
                if self._cancelado:
                    self._enviar_mensaje(
                        FaseProceso.FILTRADO,
                        NivelMensaje.WARNING,
                        "Proceso cancelado por el usuario"
                    )
                    break
                
                if idx % 100 == 0:
                    self._actualizar_progreso(idx, total_items)
                    
            except Exception as e:
                self.logger.error(f"Error en correo {idx}: {str(e)}")
                continue
        
        self.estadisticas.total_correos = len(correos_filtrados)
        self._enviar_mensaje(
            FaseProceso.FILTRADO,
            NivelMensaje.SUCCESS,
            f"Correos filtrados: {len(correos_filtrados)}"
        )
        
        return correos_filtrados
    
    def _descargar_adjuntos(self, correos_data: List[dict], carpeta, 
                           destino: str, frases: List[str]):
        """Descarga adjuntos usando Entry ID de correos filtrados con auditoría"""
        
        self._cambiar_fase(FaseProceso.DESCARGA)
        self._cambiar_estado(EstadoProceso.PROCESANDO)
        
        carpeta_path = Path(destino)
        
        # Normalizar frases para filtro de nombres de archivo
        frases_norm = [f.strip().lower() for f in frases if f.strip()]
        modo_sin_filtro = len(frases_norm) == 0
        
        total_correos = len(correos_data)
        
        for idx, correo_data in enumerate(correos_data, 1):
            try:
                # Usar Entry ID en lugar de índice para obtener correo
                entry_id = correo_data['entry_id']
                fecha_correo = correo_data['fecha']
                asunto_esperado = correo_data['asunto']
                
                # Obtener correo por Entry ID (método confiable)
                try:
                    item = self.namespace.GetItemFromID(entry_id)
                    
                    # Validación: verificar que sea el correo correcto
                    if item.Subject != asunto_esperado:
                        self.logger.warning(
                            f"⚠️ Discrepancia en correo {idx}/{total_correos}: "
                            f"Esperado '{asunto_esperado[:60]}...' "
                            f"pero obtenido '{item.Subject[:60]}...'"
                        )
                        # Continuar de todas formas, pero registrar la discrepancia
                    
                except Exception as e:
                    self.logger.error(
                        f"❌ Error obteniendo correo {idx} por Entry ID: {str(e)}"
                    )
                    self._enviar_mensaje(
                        FaseProceso.DESCARGA,
                        NivelMensaje.ERROR,
                        f"No se pudo obtener correo {idx}/{total_correos}"
                    )
                    
                    # Actualizar auditoría como error
                    self.auditor.actualizar_descarga(
                        entry_id=entry_id,
                        adjuntos_descargados=0,
                        estado_final="ERROR",
                        motivo_rechazo=f"Error obteniendo correo: {str(e)}"
                    )
                    continue
                
                adjuntos_descargados_correo = 0
                
                for adjunto in item.Attachments:
                    try:
                        nombre_archivo = adjunto.FileName
                        nombre_lower = nombre_archivo.lower()
                        
                        # Filtro por nombre de archivo (o aceptar todos si modo_sin_filtro)
                        if modo_sin_filtro:
                            coincide_nombre = True
                        else:
                            coincide_nombre = any(frase in nombre_lower for frase in frases_norm)
                        
                        if not coincide_nombre:
                            continue
                        
                        ruta_archivo = carpeta_path / nombre_archivo
                        ruta_archivo = self._manejar_nombre_duplicado(ruta_archivo)
                        
                        adjunto.SaveAsFile(str(ruta_archivo))
                        
                        tamaño_mb = ruta_archivo.stat().st_size / (1024 * 1024)
                        self.estadisticas.tamaño_total_mb += tamaño_mb
                        
                        self.estadisticas.adjuntos_descargados += 1
                        adjuntos_descargados_correo += 1
                        
                        self.estadisticas.archivos_descargados.append({
                            'nombre': ruta_archivo.name,
                            'fecha_descarga': datetime.now(),
                            'fecha_correo': fecha_correo.strftime("%d/%m/%Y"),
                            'hora_correo': fecha_correo.strftime("%H:%M:%S")
                        })
                        
                        self._enviar_mensaje(
                            FaseProceso.DESCARGA,
                            NivelMensaje.SUCCESS,
                            f"Descargado: {ruta_archivo.name}"
                        )
                    except Exception as e:
                        self.estadisticas.adjuntos_fallidos += 1
                        self.logger.error(f"Error descargando adjunto: {str(e)}")
                
                # Actualizar auditoría con resultado de descarga
                if adjuntos_descargados_correo > 0:
                    self.auditor.actualizar_descarga(
                        entry_id=entry_id,
                        adjuntos_descargados=adjuntos_descargados_correo,
                        estado_final="PROCESADO",
                        motivo_rechazo=""
                    )
                else:
                    # No se descargó ningún adjunto
                    motivo = "Adjuntos no coinciden con frases de filtrado" if not modo_sin_filtro else "Error al descargar adjuntos"
                    self.auditor.actualizar_descarga(
                        entry_id=entry_id,
                        adjuntos_descargados=0,
                        estado_final="PROCESADO",
                        motivo_rechazo=motivo
                    )
                
                self.estadisticas.correos_procesados += 1
                del item
                
            except Exception as e:
                self.logger.error(f"Error procesando correo: {str(e)}")
            
            # Verificar cancelación
            if self._cancelado:
                self._enviar_mensaje(
                    FaseProceso.DESCARGA,
                    NivelMensaje.WARNING,
                    "Proceso cancelado por el usuario"
                )
                break
            
            self._actualizar_progreso(idx, total_correos)
            
            if idx % 100 == 0:
                gc.collect()
                pythoncom.PumpWaitingMessages()
        
        self._enviar_mensaje(
            FaseProceso.DESCARGA,
            NivelMensaje.SUCCESS,
            f"Descarga completada: {self.estadisticas.adjuntos_descargados} adjuntos"
        )

    
    def _exportar_auditoria(self):
        """Exporta la auditoría a Parquet y Excel"""
        try:
            if self.auditor and len(self.auditor) > 0:
                self._enviar_mensaje(
                    FaseProceso.FINALIZACION,
                    NivelMensaje.INFO,
                    "Generando archivos de auditoría..."
                )
                
                archivos = self.auditor.exportar_todo()
                
                if archivos['excel']:
                    self._enviar_mensaje(
                        FaseProceso.FINALIZACION,
                        NivelMensaje.SUCCESS,
                        f"Auditoría exportada: {Path(archivos['excel']).name}"
                    )
                    
                    # Registrar estadísticas de auditoría
                    stats_audit = self.auditor.get_estadisticas()
                    self.logger.info(f"Estadísticas de auditoría: {stats_audit}")
                    
                    # Identificar correos problemáticos
                    problematicos = self.auditor.get_correos_problematicos()
                    if len(problematicos) > 0:
                        self.logger.warning(
                            f"⚠️ {len(problematicos)} correos con adjuntos NO descargados"
                        )
        except Exception as e:
            self.logger.error(f"Error exportando auditoría: {str(e)}")
    
    def _generar_excel_listado(self, carpeta_destino: str):
        """Genera archivo Excel con listado"""
        
        try:
            if not self.estadisticas.archivos_descargados:
                return
            
            self._enviar_mensaje(
                FaseProceso.FINALIZACION,
                NivelMensaje.INFO,
                "Generando Excel..."
            )
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Lista de Documentos"
            
            encabezados = ["Nº", "Nombre del archivo", "Fecha de descarga", 
                          "Fecha correo", "Hora correo"]
            ws.append(encabezados)
            
            # Formato encabezados
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="16A085", end_color="16A085", 
                                       fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for idx, archivo in enumerate(self.estadisticas.archivos_descargados, start=1):
                ws.append([
                    idx,
                    archivo['nombre'],
                    archivo['fecha_descarga'].strftime("%d/%m/%Y %H:%M:%S"),
                    archivo['fecha_correo'],
                    archivo['hora_correo']
                ])
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            
            # Tabla
            tabla = Table(
                displayName="TablaDocumentos",
                ref=f"A1:E{len(self.estadisticas.archivos_descargados) + 1}"
            )
            estilo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tabla.tableStyleInfo = estilo
            ws.add_table(tabla)
            
            # Guardar
            timestamp = datetime.now().strftime("%d.%m.%Y_%H.%M.%S")
            nombre_archivo = f"lista_documentos_{timestamp}.xlsx"
            ruta_excel = Path(carpeta_destino) / nombre_archivo
            
            wb.save(str(ruta_excel))
            
            self._enviar_mensaje(
                FaseProceso.FINALIZACION,
                NivelMensaje.SUCCESS,
                f"Excel generado: {nombre_archivo}"
            )
        except Exception as e:
            self.logger.error(f"Error generando Excel: {str(e)}")
    
    def _generar_resultado(self) -> dict:
        """Genera diccionario con estadísticas"""
        return {
            'total_correos': self.estadisticas.total_correos,
            'correos_procesados': self.estadisticas.correos_procesados,
            'adjuntos_descargados': self.estadisticas.adjuntos_descargados,
            'adjuntos_fallidos': self.estadisticas.adjuntos_fallidos,
            'tamaño_total_mb': self.estadisticas.tamaño_total_mb,
            'tasa_exito': self.estadisticas.tasa_exito,
            'tiempo_total': self.estadisticas.tiempo_total
        }
    
    def cancelar(self):
        """
        Marca el proceso para cancelación.
        Los loops verificarán este flag y detendrán el procesamiento.
        """
        self._cancelado = True
        self._enviar_mensaje(
            self.fase_actual,
            NivelMensaje.WARNING,
            "Cancelación solicitada..."
        )
        self.logger.warning("Proceso de extracción cancelado por el usuario")
    
    def _generar_resultado_vacio(self) -> dict:
        """Genera resultado vacío"""
        return {
            'total_correos': 0,
            'correos_procesados': 0,
            'adjuntos_descargados': 0,
            'adjuntos_fallidos': 0,
            'tamaño_total_mb': 0.0,
            'tasa_exito': 0.0,
            'tiempo_total': 0.0
        }


__all__ = [
    'ExtractorAdjuntosOutlook',
    'FaseProceso',
    'NivelMensaje',
    'EstadoProceso',
    'EstadisticasExtraccion'
]